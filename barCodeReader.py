import cv2
import os
from pyzbar import pyzbar
from openpyxl import load_workbook
from openpyxl import Workbook
import pandas as pd
import xlwt


def read_barcodes(frame):
    barcodes = pyzbar.decode(frame)
    for barcode in barcodes:
        x, y, w, h = barcode.rect
        # 1
        barcode_info = barcode.data.decode('utf-8')
        cv2.rectangle(frame, (x, y), (x+w, y+h), (0, 255, 0), 2)

        # 2
        font = cv2.FONT_HERSHEY_DUPLEX
        cv2.putText(frame, barcode_info, (x + 6, y - 6),
                    font, 2.0, (255, 255, 255), 1)
        # 3

        # print(barcode_info)
        if os.path.exists('barCodeInfo.xlsx'):
            wb = load_workbook('barCodeInfo.xlsx')
        else:
            wb = Workbook()
        sheet = wb.worksheets[0]
        if sheet['A1'].value == None:
            sheet['A1'] = '項目'
        ws = wb.active['A']
        for rowNum in range(1, len(ws)+1):
            content = sheet.cell(row=rowNum, column=1).value
            if (content != barcode_info):
                sheet.cell(row=len(ws) + 1, column=1).value = barcode_info
        wb.save('barCodeInfo.xlsx')
        wb.close()
        data = pd.DataFrame(pd.read_excel('barCodeInfo.xlsx', 'Sheet'))
        wp = data.drop_duplicates(subset=['項目'])
        wp.to_excel('barCodeInfoPure.xls')

        # with open("barcode_result.txt", mode='w') as file:
        #     file.write("Recognized Barcode:" + barcode_info)

    return frame


def main():
    # 1
    camera = cv2.VideoCapture(0)
    ret, frame = camera.read()
    # 2
    while ret:
        ret, frame = camera.read()
        frame = read_barcodes(frame)
        cv2.imshow('Barcode/QR code reader', frame)
        if cv2.waitKey(1) & 0xFF == 27:
            break
    # 3
    camera.release()
    cv2.destroyAllWindows()


# 4
if __name__ == '__main__':
    main()
